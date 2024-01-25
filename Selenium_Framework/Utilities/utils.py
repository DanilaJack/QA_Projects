import inspect
import logging
import csv
import softest
from openpyxl import workbook, load_workbook

class Utils(softest.TestCase):
    def assertListItemsText(self, list, value):
        for stop in list:
            print("The text is ", stop.text)
            self.soft_assert(self.assertEqual, stop.text, value)
            if stop.text == value:
                print("test passed")
            else:
                print("test failed")

        self.assert_all()


    def custom_logger(logLevel = logging.DEBUG):

        logger_name = inspect.stack()[1][3]
        logger = logging.getLogger(logger_name)
        logger.setLevel(logLevel)

        fh = logging.FileHandler(filename="automation.log")
        ch = logging.StreamHandler()
        formatter = logging.Formatter("%(asctime)s - %(levelname)s : %(message)s", datefmt='%Y-%m-%d %H:%M:%S')
        fh.setFormatter(formatter)
        ch.setFormatter(formatter)

        logger.addHandler(fh)
        logger.addHandler(ch)
        return logger


    def read_data_from_excel(file_name, sheet):
        data_list = []
        wb = load_workbook(filename=file_name)
        sh = wb[sheet]

        row_ct = sh.max_row
        col_ct = sh.max_column

        for i in range(2, row_ct+1):
            row = []
            for j in range(1, col_ct+1):
                row.append(sh.cell(row=i, column=j).value)
            data_list.append(row)
        return data_list

    def read_data_from_csv(filename):
        #Create an empty list
        data_list = []

        #Open csv file
        csvdata = open(filename, 'r')

        #Create a cvs reader
        reader = csv.reader(csvdata)

        #skip header
        next(reader)

        #Add csv rows to the data_list
        for rows in reader:
            data_list.append(rows)
        return data_list

