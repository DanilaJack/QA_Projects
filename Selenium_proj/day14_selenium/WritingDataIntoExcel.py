import openpyxl

# # same data in the cells
# file = "C:/Users/danil/Desktop/automated_testing/Selenium Web Automation/SeleniumPython_PavanCourse/testData.xlsx"
#
# workbook = openpyxl.load_workbook(file)
# sheet = workbook.active   # get active sheet from excel (applicable if you have only one sheet)
#
# for r in range(1,6):
#     for c in range(1,4):
#         sheet.cell(r,c).value = "welcome"
#
# workbook.save(file)


# multimple data in the cells
file = "C:/Users/danil/Desktop/automated_testing/Selenium Web Automation/SeleniumPython_PavanCourse/testData2.xlsx"

workbook = openpyxl.load_workbook(file)
sheet = workbook.active   # get active sheet from excel (applicable if you have only one sheet)

sheet.cell(1,1).value = 123
sheet.cell(1,2).value = "Smith"
sheet.cell(1,3).value = "engineer"

sheet.cell(2,1).value = 567
sheet.cell(2,2).value = "John"
sheet.cell(2,3).value = "manager"

sheet.cell(3,1).value = 987
sheet.cell(3,2).value = "David"
sheet.cell(3,3).value = "analyst"

workbook.save(file)