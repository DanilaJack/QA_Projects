import openpyxl


#File --> Workbook ---> Sheets ----> Rows ---> Cells
file = "C:/Users/danil/Desktop/automated_testing/Selenium Web Automation/SeleniumPython_PavanCourse/Userdata.xlsx"
workbook = openpyxl.load_workbook(file)
sheet = workbook["Лист1"]

rows = sheet.max_row  # count number of rows in an excel sheet
cols = sheet.max_column   # count number of columns in an excel sheet

#Reading all the cells from the excel sheet
for r in range(1, rows+1):
    for c in range(1, cols+1):
        print(sheet.cell(r,c).value, end='    ')
    print()